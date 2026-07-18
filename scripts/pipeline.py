"""Pipeline: read -> score -> enrich -> write. One module, taking a Cuisine.

Refresh of source caches is out of scope here -- run the standalone fetchers
(`step1_read_master.py`, `step2_fetch_ratings.py`, `step2b_fetch_infatuation.py`,
plus the Yelp research prompts) when caches need updating. Pipeline reads the
caches as-is.
"""
from __future__ import annotations

import json
from pathlib import Path

from scripts.scoring import ScoringConfig, score
from scripts.shared import cities, geo, places


def read(cuisine) -> list[dict]:
    return cuisine.read_restaurants()


_DEFAULT_RESOLVER = object()


def area_resolver(cuisine):
    """Resolve coordinates to a neighborhood for this cuisine's city.

    Returned as a callable so enrich() stays free of file I/O and tests can
    substitute a stub. Returns None for a cuisine whose city has no boundary
    set (the Philadelphia cuisines) — meaning "don't derive", which leaves
    their hand-entered labels untouched rather than blanking them.
    """
    region = cities.region_for(cuisine.name)
    if region is None:
        return None
    return lambda lat, lng: geo.lookup(lat, lng, region)


def score_step(cuisine, restaurants: list[dict], config: ScoringConfig | None = None) -> dict:
    cfg = config or ScoringConfig()
    adjusted = {
        s.name: {r["name"]: s.read(r["name"]) for r in restaurants}
        for s in cuisine.sources
    }
    rows = [{"name": r["name"], "price": r.get("min_price")} for r in restaurants]
    return score(rows, adjusted, cfg)


def collision_report(cuisine, restaurants: list[dict]) -> str:
    """Describe any place_id shared by two restaurants; empty when clean."""
    google = next((s for s in cuisine.sources if s.name == "google"), None)
    if google is None:
        return ""
    return places.describe_collisions(
        google.entries(), {r["name"] for r in restaurants}
    )


def trust_resolver(cuisine, restaurants: list[dict]):
    """Resolve a restaurant name to whether its Places match can be trusted.

    Returned as a callable for the same reason as area_resolver: enrich() stays
    free of I/O and tests can substitute a stub.
    """
    google = next((s for s in cuisine.sources if s.name == "google"), None)
    if google is None:
        return lambda name: False
    cache = google.entries()
    on_sheet = {r["name"] for r in restaurants}
    colliding = places.colliding_place_ids(cache, on_sheet)
    return lambda name: places.is_trustworthy(name, cache.get(name), colliding)


def enrich(cuisine, restaurants: list[dict], scored: dict, user_state: dict,
           resolve_area=_DEFAULT_RESOLVER, is_trusted=_DEFAULT_RESOLVER) -> list[dict]:
    """Compose restaurant base records with per-source readings, composite scoring,
    free-form specialty data, and preserved user_state. Pure function -- no I/O.

    resolve_area is a (lat, lng) -> Area | None callable; pass one in to keep
    this free of I/O. Defaults to the cuisine's real region resolver. Passing
    None skips neighborhood derivation entirely.

    is_trusted is a (name) -> bool callable gating anything derived from a
    Places field. Passing None treats every match as untrusted, so nothing is
    derived and hand values stand.
    """
    specialties = cuisine.load_specialties()
    resolve = area_resolver(cuisine) if resolve_area is _DEFAULT_RESOLVER else resolve_area
    trusted = (trust_resolver(cuisine, restaurants)
               if is_trusted is _DEFAULT_RESOLVER else (is_trusted or (lambda name: False)))
    enriched: list[dict] = []
    for r in restaurants:
        name = r["name"]
        rec = dict(r)
        _merge_source_fields(rec, name, cuisine.sources)
        _merge_scoring(rec, scored.get(name))
        rec.update(specialties.get(name, {}))
        _merge_user_state(rec, user_state.get(name, {}))
        _merge_geo(rec, resolve)
        _merge_closed(rec, trusted(name))
        _finalize_legacy_shape(rec)
        enriched.append(rec)
    return enriched


def _merge_closed(rec: dict, match_trusted: bool) -> None:
    """Decide whether a Restaurant is closed: explicit assertion beats Google.

    Two different things used to share the `closed` field: a deliberate "this
    place has shut" from a research file, and the pipeline's default False,
    which only ever meant "nobody said". Google's business_status can fill in
    the second without touching the first, so an assertion is preserved as
    `closed_override` and always wins.

    That ordering is load-bearing. Every restaurant where Google says
    OPERATIONAL but a research file says closed turned out to have a wrong
    Places match — the restaurant shut, Places substituted a similarly-named
    operating business (ROKI -> "RokuNana"), and its status described the
    substitute. Deriving over the override would have reopened all seven.

    Temporary closures are tracked separately: a restaurant that is dark this
    month is not the same as one that is gone, and collapsing them loses that.
    """
    override = rec.pop("closed", None)
    rec["closed_override"] = override
    status = rec.get("business_status")

    # An untrusted match's status belongs to some other business; ignore it.
    derived_permanent = status == "CLOSED_PERMANENTLY" if match_trusted else None
    derived_temporary = status == "CLOSED_TEMPORARILY" if match_trusted else None

    rec["closed"] = override if override is not None else bool(derived_permanent)
    rec["temporarily_closed"] = bool(derived_temporary) and not rec["closed"]


def _merge_geo(rec: dict, resolve) -> None:
    """Derive the neighborhood from coordinates, demoting the hand label.

    The master sheet's `neighborhood` was hand-typed and disagreed with reality
    in both directions, so it is preserved as `neighborhood_raw` for reference
    and never used for filtering. A restaurant Google could not locate keeps a
    null derived label rather than falling back to the untrustworthy one.

    With no resolver (a city we have no boundaries for), the hand label is left
    exactly where it was — deriving nothing is better than blanking the only
    label those cuisines have.
    """
    if resolve is None:
        return
    rec["neighborhood_raw"] = rec.pop("neighborhood", None)
    area = resolve(rec.get("lat"), rec.get("lng"))
    rec["borough"] = area.borough if area else None
    rec["nta_code"] = area.code if area else None
    rec["nta_name"] = area.name if area else None


_ROUND_3DP = (
    "composite_rating", "adjusted_rating", "value_score",
    "google_wilson", "yelp_wilson", "infatuation_5",
)
_PCT_FIELDS = ("rating_percentile", "value_percentile")


def _finalize_legacy_shape(rec: dict) -> None:
    for k in _ROUND_3DP:
        v = rec.get(k)
        if v is not None:
            rec[k] = round(v, 3)
    for k in _PCT_FIELDS:
        v = rec.get(k)
        if v is not None:
            rec[k] = round(v * 100, 1)
    # Specialty files use "confidence"; legacy dashboard reads "specialty_confidence".
    if "confidence" in rec:
        rec["specialty_confidence"] = rec.pop("confidence")
    # `closed` is set by _merge_closed; this only covers a cuisine that skips it.
    rec.setdefault("closed", False)


def _merge_source_fields(rec: dict, name: str, sources: list) -> None:
    for s in sources:
        entry = s.entry(name)
        reading = s.read(name)
        if s.name == "google":
            rec["raw_rating"] = entry.get("rating") if entry else None
            rec["review_count"] = entry.get("review_count") if entry else None
            rec["google_name"] = entry.get("google_name", "") if entry else ""
            rec["google_wilson"] = reading.value if reading else None
            # Places is the source of location truth; the neighborhood label is
            # derived from these downstream in _merge_geo.
            for field in ("lat", "lng", "address", "place_id", "business_status"):
                rec[field] = entry.get(field) if entry else None
        elif s.name == "yelp":
            rec["yelp_rating"] = entry.get("yelp_rating") if entry else None
            rec["yelp_count"] = entry.get("review_count") if entry else None
            rec["yelp_wilson"] = reading.value if reading else None
        elif s.name == "infatuation":
            rec["infatuation_rating"] = entry.get("rating") if entry else None
            rec["infatuation_5"] = reading.value if reading else None


def _merge_scoring(rec: dict, scoring) -> None:
    if scoring is None:
        return
    rec["composite_rating"] = scoring.composite_rating
    rec["adjusted_rating"] = scoring.composite_rating  # legacy alias
    rec["sources"] = scoring.sources
    rec["n_sources"] = scoring.sources.count("+") + 1
    rec["value_score"] = scoring.value_score
    rec["rating_percentile"] = scoring.rating_percentile
    rec["value_percentile"] = scoring.value_percentile


_USER_STATE_DEFAULTS = {
    "visited": False,
    "friend_suggested": False,
    "subway_walk_min": None,
    "nearest_456": None,
}


def _merge_user_state(rec: dict, state: dict) -> None:
    for key, default in _USER_STATE_DEFAULTS.items():
        rec[key] = state.get(key, default)


def write_scored_json(cuisine, enriched: list[dict], out_path: str | Path | None = None) -> Path:
    path = Path(out_path) if out_path else _default_data_dir(cuisine) / "scored_restaurants.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(enriched, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def write_dashboard_data(cuisine, enriched: list[dict], out_path: str | Path | None = None) -> Path:
    fields = cuisine.dashboard_fields()
    rows = [{k: r.get(k) for k in fields} for r in enriched]
    path = Path(out_path) if out_path else _default_dashboard_dir(cuisine) / "data.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
    return path


def _default_data_dir(cuisine) -> Path:
    return Path(__file__).resolve().parent / "data" / cuisine.name


def _default_dashboard_dir(cuisine) -> Path:
    return Path(__file__).resolve().parents[1] / "docs" / cuisine.name


def load_user_state(cuisine) -> dict:
    path = _default_data_dir(cuisine) / "user_state.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


_EXCEL_HEADERS = [
    ("Restaurant Name", 30), ("Neighborhood", 30), ("Price ($)", 10),
    ("Pacing", 10), ("Vibe & Distinctions", 50),
    ("Commute (Parkchester)", 12), ("Commute (Park Slope)", 12), ("Time Diff", 10),
    ("Google Rating", 12), ("Review Count", 12),
    ("Adj. Rating", 11), ("Rating Pctl", 10),
    ("Value Score", 11), ("Value Pctl", 10),
    ("Visited", 8), ("Google Maps Name", 35),
]


def write_excel(cuisine, enriched: list[dict], out_path: str | Path | None = None) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    root = Path(__file__).resolve().parents[1]
    path = Path(out_path) if out_path else root / f"{cuisine.name.capitalize()}_Ratings.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cuisine.name.capitalize()
    hfont = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    visited_fill = PatternFill("solid", fgColor="E2EFDA")
    for col, (header, width) in enumerate(_EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    for i, r in enumerate(enriched, 2):
        row_data = [
            r.get("name"),
            geo.display(r.get("borough"), r.get("nta_name"),
                        unverified_fallback=r.get("neighborhood_raw")),
            r.get("min_price"),
            r.get("pacing"), r.get("vibe"),
            r.get("commute_parkchester", ""), r.get("commute_parkslope", ""), r.get("time_diff", ""),
            r.get("raw_rating"), r.get("review_count"),
            r.get("composite_rating"), r.get("rating_percentile"),
            r.get("value_score"), r.get("value_percentile"),
            "Yes" if r.get("visited") else "", r.get("google_name", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if r.get("visited"):
                cell.fill = visited_fill
    ws.freeze_panes = "A2"
    wb.save(path)
    return path
