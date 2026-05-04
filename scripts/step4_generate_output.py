"""
Step 4: Generate the final enriched Excel spreadsheet (v2).
"""
import json, os, sys, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.insert(0, os.path.dirname(__file__))
from config import OUTPUT_EXCEL, RATING_METHOD, WILSON_Z, BAYESIAN_M, GOOGLE_BIAS_CORRECTION, PRICE_EXPONENT

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCORED_PATH = os.path.join(PROJECT_ROOT, "scripts", "scored_restaurants.json")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, OUTPUT_EXCEL)

HEADERS = [
    ("Restaurant Name", 30), ("Neighborhood", 30), ("Price ($)", 10),
    ("Pacing", 10), ("Vibe & Distinctions", 50),
    ("Commute (Parkchester)", 12), ("Commute (Park Slope)", 12), ("Time Diff", 10),
    ("Google Rating", 12), ("Review Count", 12),
    ("Adj. Rating", 11), ("Rating Pctl", 10),
    ("Value Score", 11), ("Value Pctl", 10),
    ("Visited", 8), ("Google Maps Name", 35),
]

def main():
    with open(SCORED_PATH) as f:
        restaurants = json.load(f)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Omakase"
    hfont = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    visited_fill = PatternFill("solid", fgColor="E2EFDA")
    for col, (header, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    for i, r in enumerate(restaurants, 2):
        row_data = [
            r["name"], r["neighborhood"], r["min_price"], r["pacing"], r["vibe"],
            r["commute_parkchester"], r["commute_parkslope"], r["time_diff"],
            r["raw_rating"], r["review_count"], r["adjusted_rating"],
            r.get("rating_percentile"), r["value_score"], r.get("value_percentile"),
            "Yes" if r["visited"] else "", r["google_name"],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if r["visited"]:
                cell.fill = visited_fill
    ws.freeze_panes = "A2"

    notes = wb.create_sheet("Methodology")
    notes.column_dimensions["A"].width = 90
    methodology = [
        ("Omakase Value Score Methodology (v2)", True, 14),
        ("", False, 11),
        ("1. Google Bias Correction", True, 12),
        ("   correction_factor = " + str(GOOGLE_BIAS_CORRECTION), False, 11),
        ("   corrected_rating = raw_google_rating * correction_factor", False, 11),
        ("   Google Maps ratings are systematically inflated. This corrects by ~3%.", False, 11),
        ("", False, 11),
        ("2. Wilson Score Lower Bound (Rating Adjustment)", True, 12),
        ("   z = " + str(WILSON_Z) + " (95% confidence interval)", False, 11),
        ("   Converts 5-star to proportion, computes Wilson lower bound, converts back.", False, 11),
        ("   Effect: statistically rigorous penalty for low review counts.", False, 11),
        ("   Unlike Bayesian average, penalty is based on actual statistical uncertainty.", False, 11),
        ("", False, 11),
        ("3. Price Exponent (Empirically Derived)", True, 12),
        ("   Derived via log-log regression: log(rating) ~ beta * log(price)", False, 11),
        ("   The regression coefficient |beta| becomes the exponent, capped to [0.1, 0.6].", False, 11),
        ("", False, 11),
        ("4. Value Score", True, 12),
        ("   value = adj_rating * (100 / price) ^ exponent", False, 11),
        ("   Higher score = better combination of quality and affordability.", False, 11),
        ("", False, 11),
        ("5. Percentile Ranking", True, 12),
        ("   Rating Pctl: where adjusted rating falls vs all others (0=worst, 100=best)", False, 11),
        ("   Value Pctl: where value score falls vs all others", False, 11),
        ("   Since ratings cluster 4.2-5.0, percentiles make differences interpretable.", False, 11),
    ]
    for row_num, (text, bold, size) in enumerate(methodology, 1):
        cell = notes.cell(row=row_num, column=1, value=text)
        cell.font = Font(bold=bold, size=size)

    wb.save(OUTPUT_PATH)
    scored_count = sum(1 for r in restaurants if r["value_score"] is not None)
    visited_count = sum(1 for r in restaurants if r["visited"])
    print(f"Generated {OUTPUT_PATH}")
    print(f"  {len(restaurants)} restaurants, {scored_count} with scores, {visited_count} marked visited")

if __name__ == "__main__":
    main()
